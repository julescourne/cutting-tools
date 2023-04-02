import math
import os
from controller.Switcher import Switcher
import openpyxl


class ExportExperience:
    """Classe qui permet d'importer les données d'une expérience d'usinage dans la base de données"""

    def __init__(self, id_experience):
        """Constructeur"""
        switcher = Switcher()
        switcher.experience = id_experience
        fichier_sortie = os.path.join(os.getcwd(), "nouveau_fichier.xlsx")

        workbook = openpyxl.load_workbook(
            "C:\\Users\\Jules Courné\\OneDrive\\Bureau\\Cours\\5A\\PRD\\S10\\cutting-tools\\gabarit.xlsx")

        # Sélectionner la feuille de calcul 'Sheet1'
        worksheet = workbook.get_sheet_by_name('NEW2')

        worksheet.cell(row=3, column=2).value = switcher.type_outil()[0]
        worksheet.cell(row=3, column=5).value = switcher.type_matiere_piece()[0]
        worksheet.cell(row=3, column=7).value = switcher.type_procede()[0]

        worksheet.cell(row=4, column=2).value = switcher.matiere_outil()[0]
        worksheet.cell(row=5, column=2).value = switcher.diametre_outil()[0]
        worksheet.cell(row=6, column=2).value = switcher.nb_dents_util_outil()[0]
        worksheet.cell(row=7, column=2).value = switcher.revetement_outil()[0]
        worksheet.cell(row=8, column=2).value = switcher.rayon_arrete_outil()[0]
        worksheet.cell(row=8, column=3).value = math.radians(switcher.rayon_arrete_outil()[0])

        worksheet.cell(row=9, column=2).value = switcher.angle_depouille_outil()[0]
        worksheet.cell(row=9, column=3).value = math.radians(switcher.angle_depouille_outil()[0])

        worksheet.cell(row=10, column=2).value = switcher.angle_axial_outil()[0]
        worksheet.cell(row=10, column=3).value = math.radians(switcher.angle_axial_outil()[0])

        worksheet.cell(row=11, column=2).value = switcher.angle_radial()[0]
        worksheet.cell(row=11, column=3).value = math.radians(switcher.angle_radial()[0])

        worksheet.cell(row=12, column=2).value = switcher.angle_attaque()[0]
        worksheet.cell(row=12, column=3).value = math.radians(switcher.angle_attaque()[0])

        worksheet.cell(row=13, column=2).value = switcher.angle_listel1()[0]
        worksheet.cell(row=13, column=3).value = math.radians(switcher.angle_listel1()[0])

        worksheet.cell(row=14, column=2).value = switcher.angle_listel2()[0]
        worksheet.cell(row=14, column=3).value = math.radians(switcher.angle_listel2()[0])

        worksheet.cell(row=4, column=5).value = switcher.materiaux_piece()[0]
        worksheet.cell(row=5, column=5).value = switcher.procede_elaboration()[0]
        worksheet.cell(row=6, column=5).value = switcher.impression_3d()[0]
        worksheet.cell(row=7, column=5).value = switcher.longueur_usinee()[0]
        worksheet.cell(row=8, column=5).value = switcher.num_passe()[0]

        worksheet.cell(row=4, column=7).value = switcher.type_operation()[0]
        worksheet.cell(row=5, column=7).value = switcher.assistance()[0]
        worksheet.cell(row=6, column=7).value = switcher.debit_mql()[0]
        worksheet.cell(row=7, column=7).value = switcher.debit_cryo()[0]
        worksheet.cell(row=8, column=7).value = switcher.vitesse_coupe()[0]
        worksheet.cell(row=9, column=7).value = switcher.vitesse_avance_dent()[0]
        worksheet.cell(row=10, column=7).value = switcher.profondeur_passe()[0]
        worksheet.cell(row=11, column=7).value = switcher.engagement()[0]
        worksheet.cell(row=12, column=7).value = switcher.frequence_rotation()[0]
        worksheet.cell(row=13, column=7).value = switcher.vitesse_avance_min()[0]

        cnt = 0
        for temps, fx, fy, fz in zip(switcher.temps_effort_piece(), switcher.fx_piece(), switcher.fy_piece(),
                                     switcher.fz_piece()):
            worksheet.cell(row=20 + cnt, column=1).value = temps
            worksheet.cell(row=20 + cnt, column=2).value = fx
            worksheet.cell(row=20 + cnt, column=3).value = fy
            worksheet.cell(row=20 + cnt, column=4).value = fz
            cnt += 1

        cnt = 0
        for temps, tempera in zip(switcher.temps_temperature_piece(), switcher.temperature_piece()):
            worksheet.cell(row=20 + cnt, column=5).value = temps
            worksheet.cell(row=20 + cnt, column=6).value = tempera
            cnt += 1

        cnt = 0
        for contr in switcher.contrainte_residuelle():
            worksheet.cell(row=20 + cnt, column=7).value = contr
            cnt += 1

        cnt = 0
        for lim in switcher.limite_endurance():
            worksheet.cell(row=20 + cnt, column=8).value = lim
            cnt += 1

        cnt = 0
        for dure in switcher.durete():
            worksheet.cell(row=20 + cnt, column=9).value = dure
            cnt += 1

        cnt = 0
        for rug in switcher.rugosite():
            worksheet.cell(row=20 + cnt, column=10).value = rug
            cnt += 1

        cnt = 0
        for tmp in switcher.temps_usinage():
            worksheet.cell(row=20 + cnt, column=11).value = tmp
            cnt += 1

        cnt = 0
        for vb in switcher.vb():
            worksheet.cell(row=20 + cnt, column=12).value = vb
            cnt += 1

        cnt = 0
        for er in switcher.er():
            worksheet.cell(row=20 + cnt, column=13).value = er
            cnt += 1

        cnt = 0
        for kt in switcher.kt():
            worksheet.cell(row=20 + cnt, column=14).value = kt
            cnt += 1

        cnt = 0
        for temps, tempera in zip(switcher.temps_temperature_outil(), switcher.temperature_outil()):
            worksheet.cell(row=20 + cnt, column=15).value = temps
            worksheet.cell(row=20 + cnt, column=16).value = tempera
            cnt += 1

        cnt = 0
        for ep in switcher.epaisseur_copeaux():
            worksheet.cell(row=20 + cnt, column=17).value = ep
            cnt += 1

        cnt = 0
        for tmp in switcher.temps_vibration():
            worksheet.cell(row=20 + cnt, column=18).value = tmp
            cnt += 1

        cnt = 0
        for freq in switcher.frequence():
            worksheet.cell(row=20 + cnt, column=19).value = freq
            cnt += 1

        cnt = 0
        for amp in switcher.amplitude():
            worksheet.cell(row=20 + cnt, column=20).value = amp
            cnt += 1

        name = ''
        if switcher.procede_elaboration()[0] == 'Conventionnel':
            name += 'CONV_'
        elif switcher.procede_elaboration()[0] == 'Imprimé 3D':
            name += str(switcher.impression_3d()[0]) + '_'

        name += str(switcher.materiaux_piece()[0]) + '_Vc_' + str(switcher.vitesse_coupe()[0]) + '_fz_' + \
                str(switcher.vitesse_avance_dent()[0]) + '_' + str(switcher.assistance()[0]) + '.xlsx'

        # Enregistrer les modifications dans le fichier Excel existant
        workbook.save(name)
